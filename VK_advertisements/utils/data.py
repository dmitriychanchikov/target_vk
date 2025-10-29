import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


# Input data

def check_dir(dir):
    try:
        if not os.path.exists(dir) and not os.path.isdir(dir):
            print(f"Директория {dir} не существует или не является директории")
            return False
        print(f"Директория '{dir}' найдена")
        return True
    except Exception as e:
        print(f"Ошибка при поиске директории '{dir}': {str(e)}")
        return None
    
def find_excel(input_dir, excel_name='names'):
    try:
        excel_path = f'{input_dir}/{excel_name}.xlsx'
        if not os.path.exists(excel_path):
            print(f"Файл '{excel_name}.xlsx' не найден в директории {input_dir}")
            return None
        print(f"Файл '{excel_name}.xlsx' найден")
        return excel_path
    except Exception as e:
        print(f"Ошибка при поиске файла '{excel_name}.xlsx': {str(e)}")
        return None
        
def load_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        values, links = [], []

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value is None or cell.hyperlink is None:
                continue
            values.append(cell.value)
            links.append(cell.hyperlink.target)

        df = pd.DataFrame({'url': links, 'name': values})
        if df.empty:
            print(f"В файле '{excel_path}' не обнаружны записи со ссылками")
            return None
        print(f"Получено {len(df)} записей из файла '{excel_path}'")
        return df
    
    except Exception as e:
        print(f"Ошибка при загрузке файла '{excel_path}': {str(e)}")
        return None
    
def find_photo_dir(input_dir, dir_name='photos'):
    try:
        photo_dir = f'{input_dir}/{dir_name}'
        if not os.path.exists(photo_dir) and not os.path.isdir(photo_dir):
            print(f"Директория '{photo_dir}' не существует или путь не является директорией")
            return None
        print(f"Директория '{photo_dir}' найдена")
        return photo_dir
    except Exception as e:
        print(f"Ошибка при поиске директории с фотографиями: {str(e)}")
        return None
    
def find_photo(photo_dir, name):
    try:
        if not os.path.exists(photo_dir) or not os.path.isdir(photo_dir):
            print(f"Директория '{photo_dir}' не существует или путь не является директорией")
            return None
        
        for file in os.listdir(photo_dir):
            try:
                file_path = f'{photo_dir}/{file}'
                if not os.path.isfile(file_path):
                    continue
                file_name, file_ext = os.path.splitext(file)
                if file_ext.lower() not in ['.jpg', 'jpeg', '.png', '.gif', '.heic', '.heif', '.webp']:
                    continue
                if file_name == name:
                    return os.path.abspath(file_path)                
            except PermissionError:
                continue            
            except Exception as exc:
                print(f"Ошибка при обработке файла '{file}': {str(exc)}")
                continue
        
        print (f"Файл, содержащий '{name}' и имеющий расширение изображения, не найден в директории '{photo_dir}'")
        return None
    
    except Exception as e:
        print(f"Ошибка при поиске фотографии: {str(e)}")
        return None
    
def find_sign(sign_path):
    try:
        if not os.path.exists(sign_path) or not os.path.isfile(sign_path):
            print(f"Файл '{sign_path}' не существует или путь не является файлом")
            return None

        _, file_ext = os.path.splitext(sign_path)
        if file_ext.lower() not in ['.jpg', '.jpeg', '.png', '.gif', '.heic', '.heif', '.webp']:
            print(f"Файл '{sign_path}' имеет недопустимое расширение для изображения")
            return None

        print (f"Знак '{sign_path}' найден")
        return os.path.abspath(sign_path)
    
    except Exception as e:
        print(f"Ошибка при поиске знака: {str(e)}")
        return None
    
def process_input_data(input_dir, sign_path, message, pattern):
    try:
        if not check_dir(input_dir):
            return None
        
        excel_path = find_excel(input_dir)
        if excel_path is None:
            return None
        
        input_df = load_excel(excel_path)
        if input_df is None:
            return None
        
        sign_abs_path = find_sign(sign_path)
        if sign_abs_path is None:
            return None
        
        photo_dir = find_photo_dir(input_dir)
        if photo_dir is None:
            return None
        
        photo_abs_paths = []
        for _, row in input_df.iterrows():
            name = row['name']
            photo_abs_path = find_photo(photo_dir, name)
            photo_abs_paths.append(photo_abs_path)
        
        input_df['photo_path'] = photo_abs_paths
        input_df['sign_path'] = [sign_abs_path] * len(input_df)
        input_df['message'] = [message] * len(input_df)
        input_df['pattern'] = [pattern] * len(input_df)
        return input_df
    
    except Exception as e:
        print(f"Ошибка при обработке входных данных: {str(e)}")
        return None


# Output data

def init_output_dir(input_dir: str):
    """
    Создает выходную директорию и возвращает пути к CSV файлам
    """
    try:
        output_dir = input_dir.replace('in', 'out')
        os.makedirs(output_dir, exist_ok=True)
        success_file = os.path.join(output_dir, 'success.csv')
        failure_file = os.path.join(output_dir, 'failure.csv')
        return success_file, failure_file

    except Exception as e:
        print(f"Ошибка при создании выходной директории: {str(e)}")
        return False

def save_dict_to_csv(data: dict, filename: str):
    """
    Сохраняет одну запись (словарь) в CSV.
    Если файл пустой или еще не существует, добавляет заголовки.
    """
    try:
        df = pd.DataFrame([data])  # превращаем словарь в dataframe из одной строки
        file_exists = os.path.isfile(filename)

        df.to_csv(filename, mode='a', index=False, header=not file_exists, encoding='utf-8-sig')
        return True

    except Exception as e:
        print(f"Ошибка при сохранении словаря в {filename}: {e}")
        return False

def delete_dict_from_csv(data: dict, filename: str, key_field: str = "url"):
    """
    Удаляет строку из CSV по значению ключевого поля (по умолчанию 'url').
    """
    try:
        if not os.path.exists(filename):
            print(f"Файл {filename} не найден")
            return False

        df = pd.read_csv(filename, encoding='utf-8-sig')

        # фильтруем — оставляем только те строки, у которых ключ не совпадает
        initial_len = len(df)
        df = df[df[key_field] != data[key_field]]

        if len(df) == initial_len:
            print(f"Запись с {key_field}={data[key_field]} не найдена")
            return False

        # сохраняем обратно
        df.to_csv(filename, index=False, encoding='utf-8-sig')
        print(f"Запись с {key_field}={data[key_field]} удалена")
        return True

    except Exception as e:
        print(f"Ошибка при удалении записи из {filename}: {e}")
        return False

def get_df_from_csv(filename: str):
    """
    Читает CSV и возвращает pandas DataFrame.
    """
    try:
        if not os.path.exists(filename):
            print(f"Файл {filename} не найден")
            return None

        return pd.read_csv(filename, encoding='utf-8-sig')

    except Exception as e:
        print(f"Ошибка при чтении {filename}: {e}")
        return None
