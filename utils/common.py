import os
import time
import json
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import (
    TimeoutException, 
    WebDriverException
)
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from settings.config import SLEEP_TIME, WAIT_TIME


# Input data

def check_dir(dir):
    try:
        if not os.path.exists(dir) and not os.path.isdir(dir):
            print(f"Директория {dir} не существует или не является директории")
            return False
        print(f"Директория '{dir}' найдена")
        return True
    except Exception as e:
        print(f"Ошибка при поиске директории '{dir}': {str(e)}")
        return None
    
def find_excel(input_dir, excel_name='names'):
    try:
        excel_path = f'{input_dir}/{excel_name}.xlsx'
        if not os.path.exists(excel_path):
            print(f"Файл '{excel_name}.xlsx' не найден в директории {input_dir}")
            return None
        print(f"Файл '{excel_name}.xlsx' найден")
        return excel_path
    except Exception as e:
        print(f"Ошибка при поиске файла '{excel_name}.xlsx': {str(e)}")
        return None
        
def load_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        values, links = [], []

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value is None or cell.hyperlink is None:
                continue
            values.append(cell.value)
            links.append(cell.hyperlink.target)

        df = pd.DataFrame({'url': links, 'name': values})
        if df.empty:
            print(f"В файле '{excel_path}' не обнаружны записи со ссылками")
            return None
        print(f"Получено {len(df)} записей из файла '{excel_path}'")
        return df
    
    except Exception as e:
        print(f"Ошибка при загрузке файла '{excel_path}': {str(e)}")
        return None
    
def find_photo_dir(input_dir, dir_name='photos'):
    try:
        photo_dir = f'{input_dir}/{dir_name}'
        if not os.path.exists(photo_dir) and not os.path.isdir(photo_dir):
            print(f"Директория '{photo_dir}' не существует или путь не является директорией")
            return None
        print(f"Директория '{photo_dir}' найдена")
        return photo_dir
    except Exception as e:
        print(f"Ошибка при поиске директории с фотографиями: {str(e)}")
        return None
    
def find_photo(photo_dir, name):
    try:
        if not os.path.exists(photo_dir) or not os.path.isdir(photo_dir):
            print(f"Директория '{photo_dir}' не существует или путь не является директорией")
            return None
        
        for file in os.listdir(photo_dir):
            try:
                file_path = f'{photo_dir}/{file}'
                if not os.path.isfile(file_path):
                    continue
                file_name, file_ext = os.path.splitext(file)
                if file_ext.lower() not in ['.jpg', 'jpeg', '.png', '.gif', '.heic', '.heif', '.webp']:
                    continue
                if file_name == name:
                    return os.path.abspath(file_path)                
            except PermissionError:
                continue            
            except Exception as exc:
                print(f"Ошибка при обработке файла '{file}': {str(exc)}")
                continue
        
        print (f"Файл, содержащий '{name}' и имеющий расширение изображения, не найден в директории '{photo_dir}'")
        return None
    
    except Exception as e:
        print(f"Ошибка при поиске фотографии: {str(e)}")
        return None
    
def find_sign(sign_path):
    try:
        if not os.path.exists(sign_path) or not os.path.isfile(sign_path):
            print(f"Файл '{sign_path}' не существует или путь не является файлом")
            return None

        _, file_ext = os.path.splitext(sign_path)
        if file_ext.lower() not in ['.jpg', '.jpeg', '.png', '.gif', '.heic', '.heif', '.webp']:
            print(f"Файл '{sign_path}' имеет недопустимое расширение для изображения")
            return None

        print (f"Знак '{sign_path}' найден")
        return os.path.abspath(sign_path)
    
    except Exception as e:
        print(f"Ошибка при поиске знака: {str(e)}")
        return None
    
def process_input_data(input_dir, sign_path, message, pattern):
    try:
        if not check_dir(input_dir):
            return None
        
        excel_path = find_excel(input_dir)
        if excel_path is None:
            return None
        
        input_df = load_excel(excel_path)
        if input_df is None:
            return None
        
        sign_abs_path = find_sign(sign_path)
        if sign_abs_path is None:
            return None
        
        photo_dir = find_photo_dir(input_dir)
        if photo_dir is None:
            return None
        
        photo_abs_paths = []
        for _, row in input_df.iterrows():
            name = row['name']
            photo_abs_path = find_photo(photo_dir, name)
            photo_abs_paths.append(photo_abs_path)
        
        input_df['photo_path'] = photo_abs_paths
        input_df['sign_path'] = [sign_abs_path] * len(input_df)
        input_df['message'] = [message] * len(input_df)
        input_df['pattern'] = [pattern] * len(input_df)
        return input_df
    
    except Exception as e:
        print(f"Ошибка при обработке входных данных: {str(e)}")
        return None


# Output data

def init_output_dir(input_dir):
    """Создает выходную директорию и возвращает пути к файлам"""
    output_dir = input_dir.replace('in', 'out')
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%H_%M_%S')
    return (
        f'{output_dir}/{timestamp}__success.json',
        f'{output_dir}/{timestamp}__failure.json'
    )

def save_dict_to_json(data, filename):
    """
    Сохраняет данные в JSON файл в правильном формате.
    Если файл существует - загружает его, обновляет данные и сохраняет.
    Если нет - создает новый с данными в списке.
    """
    try:
        # Переводим словарь в JSON строку и добавляем новую строку
        with open(filename, 'a', encoding='utf-8') as f:
            json_string = json.dumps(data, ensure_ascii=False)
            f.write(json_string + '\n')
    except (FileNotFoundError, json.JSONDecodeError):
        # Если файла нет или он пустой/невалидный
        raise Exception(f"Невозможно сохранить данные в файл {filename}")

def get_df_from_json(filename):
    """
    Читает JSON Lines файл и возвращает pandas DataFrame.
    """
    try:
        data = []
        # Чтение файла построчно
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():  # Если строка не пустая
                    data.append(json.loads(line))
        return pd.DataFrame(data)
    except (FileNotFoundError, json.JSONDecodeError):
        # Если файла нет или он пустой/невалидный
        raise Exception(f"Невозможно прочитать файл {filename}")


# Browser

def load_webdriver():
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service)
        driver.maximize_window()
        print("Браузер запущен")
        time.sleep(SLEEP_TIME)
        return driver
    except WebDriverException as e:
        print(f"Не удалось запустить браузер: {str(e)}")
        return None
    except Exception as e:
        print(f"Неожиданная ошибка при запуске браузера: {str(e)}")
        return None

def load_website(driver, url):
    try:
        driver.get(url)
        WebDriverWait(driver, WAIT_TIME).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        print(f"Страница '{url}' загружена")
        time.sleep(SLEEP_TIME)
        return True
    except TimeoutException:
        print(f"Страница '{url}' не загрузилась за {WAIT_TIME} секунд")
        return True
    except WebDriverException as e:
        print(f"Ошибка при загрузке страницы: {str(e)}")
        return True
    except Exception as e:
        print(f"Неожиданная ошибка при загрузке страницы: {str(e)}")
        return True

def reload_website(driver):
    try:
        driver.refresh()
        WebDriverWait(driver, WAIT_TIME).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        print(f"Страница перезагружена")
        time.sleep(SLEEP_TIME)
        return True
    except TimeoutException:
        print(f"Страница не перезагрузилась за {WAIT_TIME} секунд")
        return False
    except WebDriverException as e:
        print(f"Ошибка при загрузке страницы: {str(e)}")
        return False
    except Exception as e:
        print(f"Неожиданная ошибка при перезагрузке страницы: {str(e)}")
        return False
    
def check_and_get_new_url(driver, url):
    try:
        WebDriverWait(driver, WAIT_TIME).until(
            lambda d: d.current_url != url
        )
        print(f"Загружена новая страница: '{driver.current_url}'")
        return driver.current_url
    except TimeoutException:
        print(f"URL не изменился, новая страница не загрузилась за {WAIT_TIME} секунд")
        return None
    except WebDriverException as e:
        print(f"Ошибка при загрузке страницы: {str(e)}")
        return None
    except Exception as e:
        print(f"Неожиданная ошибка при загрузке страницы: {str(e)}")
        return None
    
def shut_webdriver(driver):
    try:
        time.sleep(WAIT_TIME)
        driver.quit()
        print("Браузер закрыт")
        time.sleep(SLEEP_TIME)
        return True
    except WebDriverException as e:
        print(f"Не удалось остановить браузер: {str(e)}")
        return False
    except Exception as e:
        print(f"Неожиданная ошибка при остановке браузера: {str(e)}")
        return False


# Common actions

def click_safely(element):
    try:
        element.click()
        time.sleep(SLEEP_TIME)
        return True
    except Exception as e:
        print(f"Ошибка клика по элементу: {str(e)}")
        return False

def insert_value_safely(driver, element, value):
    try:
        driver.execute_script(f"arguments[0].value = `{value}`;", element)
        element.send_keys(" ", Keys.BACKSPACE)
        return True
    except Exception as e:
        print(f"Ошибка вставки значения: {str(e)}")
        return False
    
def replace_html_safely(driver, element, new_html):
    try:
        driver.execute_script(f"arguments[0].innerHTML = `{new_html}`;", element)
        driver.execute_script("""
            var event = new Event('input', { bubbles: true });
            arguments[0].dispatchEvent(event);
        """, element)
        return True
    except Exception as e:
        print(f"Ошибка замены текста: {str(e)}")
        return False

def press_escape(driver):
    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(SLEEP_TIME)
        return True
    except Exception as e:
        print(f"Ошибка при нажатии клавиши ESCAPE: {str(e)}")
        return False

# def click_by_coordinates(driver, x, y):
#     try:
#         actions = ActionChains(driver)
#         actions.move_by_offset(x, y).click().perform()
#         actions.move_by_offset(-x, -y).perform()
#         time.sleep(SLEEP_TIME)
#         return True
#     except Exception as e:
#         print(f"Ошибка при клике по координатам ({x}, {y}): {str(e)}")
#         return False
